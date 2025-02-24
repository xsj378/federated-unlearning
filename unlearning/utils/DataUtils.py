import os
from openpyxl import Workbook, load_workbook
import pickle
from utils import Arguments

Arg = Arguments.Arg()

class DataUtils():
    def __init__(self):
        self.path = Arg.save_path

    def insert_data_to_excel(self,excel_file, row_data, sheet_name="Sheet"):
        """
        在 Excel 中插入一行数据，如果 Excel 文件不存在就创建一个

        :param excel_file: Excel 文件名
        :param sheet_name: 工作表名
        :param row_data: 要插入的一行数据，以列表形式表示
        :return: 无
        """
        save_path = self.path + "/excels/"
        excel_file = save_path + excel_file
        # 如果 Excel 文件不存在，则创建一个
        if not os.path.exists(excel_file):
            workbook = Workbook()
            workbook.save(filename=excel_file)

        # 打开 Excel 文件并选择工作表
        workbook = load_workbook(filename=excel_file)
        worksheet = workbook[sheet_name]

        # 在工作表中插入一行数据
        worksheet.append(row_data)

        # 保存 Excel 文件
        workbook.save(filename=excel_file)

    def save_var(self,name,data):
        path = self.path + "/vars/"
        var_path = path + name
        with open(var_path, 'wb') as f:
            pickle.dump(data, f)

    def load_var(self,name):
        path = self.path + "/vars/"
        var_path = path + name
        with open(var_path, 'rb') as f:
            data = pickle.load(f)
        return data

class Global_Variable:
    def __init__(self, usersNumber):
        self.variables = {}
        self.usersNumber = usersNumber

    def insert_var(self, var_name, data, format = "static", epoch = False):
        """
        插入list类型的数据就是元素值append值
        如果 epoch 为 True 说明 该变量是每回合都会存储因此是一个矩阵
        :param format: 数据格式
        :param var_name: 数据格式
        :param data: 数据
        :return:
        """
        if format == 'list':
            if epoch == False:
                if var_name in self.variables:
                    self.variables[var_name].append(data)
                else:
                    self.variables[var_name] = [data]
            else:
                if var_name in self.variables:
                    if(len(self.variables[var_name][-1]) == self.usersNumber):
                        self.variables[var_name].append([data])
                    else: self.variables[var_name][-1].append(data)
                else:
                    self.variables[var_name] = [[data]]
        elif format == 'static':
            self.variables[var_name] = data

    def get_var(self, var_name):
        return self.variables.get(var_name)

    def print_var(self, var_name):
        print(self.variables.get(var_name))

    def show_all(self):
        print(self.variables.keys())
        for key in self.variables.keys():
            print("key:{},data:{}".format(key,self.variables[key]))